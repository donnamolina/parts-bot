"""
Excel report generator for parts comparison.
DCS blue (#003DA5) headers, clickable links, color-coded savings,
landed cost with ClickPack courier, summary row.
"""

import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Colors
DCS_BLUE = "003DA5"
MEDIUM_BLUE = "2E75B6"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
LIGHT_BLUE = "EBF3FB"
GREEN_BG = "E2EFDA"
YELLOW_BG = "FFF2CC"
ORANGE_BG = "FCE4D6"

# Fonts
GREEN_FONT = Font(color="008000", bold=True, size=10, name="Calibri")
RED_FONT = Font(color="CC0000", bold=True, size=10, name="Calibri")
LINK_FONT = Font(color="0563C1", underline="single", size=9, name="Calibri")

# Border
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(cell, bg=DCS_BLUE):
    cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def _style_data(cell, bg=WHITE, bold=False, align="left"):
    cell.font = Font(bold=bold, color="000000", size=10, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER


def _confidence_label(result: dict, is_flagged: bool) -> str:
    """Compute confidence label for a single result row."""
    if is_flagged:
        return "🔴 Revisar"
    if result.get("from_cache"):
        return "✅ Verificado"
    best = result.get("best_option") or {}
    if not best.get("price"):
        return "🔴 Revisar"
    # Cross-platform OEM# from RockAuto — wrong car's catalog
    if result.get("oem_platform_mismatch"):
        return "🔴 Verificar plataforma"
    # 7zap VIN-exact match
    oem_src = result.get("oem_source", "")
    oem_conf = result.get("oem_confidence", "")
    if oem_src == "7zap_vin_exact":
        return "🟢 7zap VIN"
    if oem_src == "7zap_fuzzy":
        return "🟡 7zap verificar"
    # RockAuto / legacy — confidence based on OEM# presence
    pn = best.get("part_number") or ""
    import re as _re
    is_real_oem = bool(pn and _re.search(r'\d', pn) and 5 <= len(pn) <= 18 and not pn.isalpha())
    if is_real_oem:
        return "🟢 Alto"
    if best.get("price", 0) > 0:
        return "🟡 Medio"
    return "🔴 Revisar"


def _build_scenario_sheet(wb, vehicle_info: dict, main_sr: int, main_sr_sup: int,
                          total_landed: float, supplier_total_dop: float | None):
    """Add 'Escenario %' worksheet with markup scenarios and supplier comparison."""
    MAIN = "'Comparacion de Precios'"
    landed_ref = f"={MAIN}!{get_column_letter(11)}{main_sr}"   # main sheet total landed cell
    supplier_col = get_column_letter(11)                         # K = Costo Total column

    year = vehicle_info.get("year", "")
    make = vehicle_info.get("make", "")
    model = vehicle_info.get("model", "")

    ws = wb.create_sheet(title="Escenario %")
    ws.sheet_view.zoomScale = 100

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18

    # ── Row 1: Vehicle header ──
    ws.merge_cells("A1:E1")
    hdr = ws["A1"]
    hdr.value = f"Escenario de Precios — {year} {make} {model}"
    hdr.font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    hdr.fill = PatternFill("solid", fgColor=DCS_BLUE)
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Row 2: empty ──

    # ── Section 1: Markup table ──
    # Row 3: headers
    table_headers = ["Markup %", "Costo Total (RD$)", "Precio de Venta (RD$)", "Ganancia (RD$)"]
    for col, h in enumerate(table_headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        _style_header(c)
    ws.row_dimensions[3].height = 25

    # Rows 4–9: 5% through 30%
    markups = [5, 10, 15, 20, 25, 30]
    for offset, pct in enumerate(markups):
        row = 4 + offset
        bg = LIGHT_BLUE if offset % 2 == 0 else WHITE

        # Markup % — stored as integer (e.g. 5 = 5%), displayed as "5%"
        a = ws.cell(row=row, column=1, value=pct)
        a.number_format = '0"%"'
        _style_data(a, bg=bg, align="center", bold=True)

        # Costo total — row 4 references main sheet; rows 5-9 reference B4
        b = ws.cell(row=row, column=2,
                    value=landed_ref if row == 4 else "=B4")
        b.number_format = "#,##0"
        _style_data(b, bg=bg, align="center")

        # Precio de venta = Costo × (1 + markup/100)
        c = ws.cell(row=row, column=3, value=f"=B{row}*(1+A{row}/100)")
        c.number_format = "#,##0"
        _style_data(c, bg=bg, align="center", bold=True)

        # Ganancia = Precio - Costo
        d = ws.cell(row=row, column=4, value=f"=C{row}-B{row}")
        d.number_format = "#,##0"
        _style_data(d, bg=bg, align="center")
        d.font = Font(color="008000", bold=False, size=10, name="Calibri")

        ws.row_dimensions[row].height = 20

    # ── Row 10: empty separator ──

    # ── Section 2: Custom markup ──
    # Row 11: section header
    ws.merge_cells("A11:E11")
    sec2 = ws["A11"]
    sec2.value = "Margen personalizado"
    _style_header(sec2, bg=MEDIUM_BLUE)
    ws.row_dimensions[11].height = 22

    # Row 12: label + yellow input cell + % sign
    ws.cell(row=12, column=1, value="Escribe el porcentaje (ej: 20):")
    _style_data(ws.cell(row=12, column=1))

    input_cell = ws.cell(row=12, column=2)
    input_cell.fill = PatternFill("solid", fgColor=YELLOW_BG)
    input_cell.border = Border(
        left=Side(border_style="medium", color="FFA500"),
        right=Side(border_style="medium", color="FFA500"),
        top=Side(border_style="medium", color="FFA500"),
        bottom=Side(border_style="medium", color="FFA500"),
    )
    input_cell.alignment = Alignment(horizontal="center", vertical="center")
    input_cell.font = Font(bold=True, size=12, name="Calibri")

    ws.cell(row=12, column=3, value="%")
    _style_data(ws.cell(row=12, column=3))
    ws.row_dimensions[12].height = 22

    # Row 13: calculated results
    ws.cell(row=13, column=1, value="Precio de Venta:")
    _style_data(ws.cell(row=13, column=1))

    pv = ws.cell(row=13, column=2, value='=IF(B12="","",B4*(1+B12/100))')
    pv.number_format = "#,##0"
    _style_data(pv, align="center", bold=True)

    ws.cell(row=13, column=3, value="Ganancia:")
    _style_data(ws.cell(row=13, column=3))

    gan = ws.cell(row=13, column=4, value='=IF(B12="","",B4*(B12/100))')
    gan.number_format = "#,##0"
    _style_data(gan, align="center")
    gan.font = Font(color="008000", bold=True, size=10, name="Calibri")

    ws.row_dimensions[13].height = 20

    # ── Row 14: empty separator ──

    # ── Section 3: Supplier comparison ──
    # Row 15: section header
    ws.merge_cells("A15:E15")
    sec3 = ws["A15"]
    sec3.value = "Comparación con cotización del suplidor"
    _style_header(sec3, bg=MEDIUM_BLUE)
    ws.row_dimensions[15].height = 22

    # Row 16: supplier total
    ws.cell(row=16, column=1, value="Cotización del suplidor (RD$):")
    _style_data(ws.cell(row=16, column=1))

    if supplier_total_dop and supplier_total_dop > 0:
        sup_cell = ws.cell(row=16, column=2,
                           value=f"={MAIN}!{supplier_col}{main_sr_sup}")
    else:
        sup_cell = ws.cell(row=16, column=2, value="N/D")
    sup_cell.number_format = "#,##0"
    _style_data(sup_cell, align="center", bold=True)
    ws.row_dimensions[16].height = 20

    # Row 17: profit vs supplier
    ws.cell(row=17, column=1, value="Si vendemos al precio suplidor, ganancia:")
    _style_data(ws.cell(row=17, column=1))

    if supplier_total_dop and supplier_total_dop > 0:
        profit = ws.cell(row=17, column=2, value='=IF(B16="N/D","N/D",B16-B4)')
        profit.number_format = "#,##0"
        _style_data(profit, align="center")
        profit.font = Font(color="008000", bold=True, size=10, name="Calibri")

        pct_cell = ws.cell(row=17, column=3,
                           value='=IF(OR(B16="N/D",B16=0),"N/D",(B16-B4)/B16)')
        pct_cell.number_format = "0.0%"
        _style_data(pct_cell, align="center")
        pct_cell.font = Font(color="008000", bold=True, size=10, name="Calibri")
    else:
        nd = ws.cell(row=17, column=2, value="N/D — no hay cotización del suplidor")
        _style_data(nd)

    ws.row_dimensions[17].height = 20


def generate_excel(results: list, vehicle_info: dict, output_path: str,
                   supplier_total_dop: float = None,
                   sonnet_flags: list = None) -> str:
    """Generate comparison Excel from search results.

    Args:
        results: List of result dicts from search_all_parts()
        vehicle_info: Dict with vin, year, make, model
        output_path: Where to save the .xlsx file

    Returns:
        Path to generated file.
    """
    exchange_rate = float(os.getenv("EXCHANGE_RATE_DOP_USD", "63"))
    clickpack_rate = float(os.getenv("CLICKPACK_RATE_DOP_PER_LB", "246"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparacion de Precios"
    ws.sheet_view.zoomScale = 90

    # ── Vehicle info row ──
    vin = vehicle_info.get("vin", "N/A")
    year = vehicle_info.get("year", "")
    make = vehicle_info.get("make", "")
    model = vehicle_info.get("model", "")

    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value = (f"Vehiculo: {year} {make} {model} | VIN: {vin} | "
                        f"Fecha: {datetime.now().strftime('%d/%b/%Y %I:%M %p')}")
    title_cell.font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor=DCS_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Courier info row
    ws.merge_cells("A2:R2")
    info_cell = ws["A2"]
    info_cell.value = (f"Courier: ClickPack | Tarifa: RD${clickpack_rate:.0f}/lb | "
                       f"Cambio: RD${exchange_rate:.0f} = $1 USD")
    info_cell.font = Font(size=9, color="444444", name="Calibri")
    info_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    info_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Parse flagged part indices from sonnet_flags (format: "#4: issue text")
    import re as _re
    flagged_indices = set()
    for flag in (sonnet_flags or []):
        m = _re.search(r'#(\d+)', flag)
        if m:
            flagged_indices.add(int(m.group(1)) - 1)  # 0-based

    # ── Column headers (row 3) ──
    headers = [
        "#", "Pieza (Original)", "Pieza (EN)", "Cant.", "Lado", "OEM #",
        "Precio (USD)", "Envio US (USD)",
        "Peso (lbs)", "Courier (RD$)", "Costo Total (RD$)",
        "Entrega (días)", "Confianza", "Fuente", "Link",
    ]
    col_widths = [4, 22, 20, 6, 8, 16, 13, 13, 10, 13, 16, 12, 14, 10, 40]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=header)
        _style_header(cell)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "A4"

    # ── Data rows (starting row 4) ──
    total_landed = 0
    parts_found = 0
    max_delivery_days = 0
    any_delivery_known = False

    for i, result in enumerate(results, 1):
        row = i + 3
        part = result.get("part", {})
        best = result.get("best_option") or {}
        landed = result.get("landed_cost") or {}

        landed_dop = landed.get("total_landed_dop", 0) if landed else 0
        # qty-adjusted totals are handled inside the column block below

        bg = LIGHT_BLUE if i % 2 == 0 else WHITE

        qty = part.get("quantity", 1) or 1

        # Column data  (cols shifted +1 vs old layout due to new Cant. col)
        ws.cell(row=row, column=1, value=i)
        _style_data(ws.cell(row=row, column=1), bg=bg, align="center")

        ws.cell(row=row, column=2, value=part.get("name_original", ""))
        _style_data(ws.cell(row=row, column=2), bg=bg)

        ws.cell(row=row, column=3, value=part.get("name_english", ""))
        _style_data(ws.cell(row=row, column=3), bg=bg)

        # Cantidad
        ws.cell(row=row, column=4, value=qty)
        _style_data(ws.cell(row=row, column=4), bg=bg, align="center", bold=(qty > 1))

        side_display = (part.get("side") or "").upper() or "N/A"
        pos = part.get("position") or ""
        if pos:
            side_display = f"{pos.title()} {side_display}" if side_display != "N/A" else pos.title()
        ws.cell(row=row, column=5, value=side_display)
        _style_data(ws.cell(row=row, column=5), bg=bg, align="center")

        ws.cell(row=row, column=6, value=best.get("part_number") or "N/F")
        _style_data(ws.cell(row=row, column=6), bg=bg, align="center")

        # USD price (unit price from listing)
        if best.get("price"):
            ws.cell(row=row, column=7, value=round(best["price"], 2))
            _style_data(ws.cell(row=row, column=7), bg=bg, align="center", bold=True)
            ws.cell(row=row, column=7).number_format = '$#,##0.00'
        else:
            ws.cell(row=row, column=7, value="N/F")
            _style_data(ws.cell(row=row, column=7), bg=YELLOW_BG, align="center")

        # US Shipping
        if landed:
            ws.cell(row=row, column=8, value=landed.get("us_shipping_usd", 0))
            _style_data(ws.cell(row=row, column=8), bg=bg, align="center")
            ws.cell(row=row, column=8).number_format = '$#,##0.00'
        else:
            ws.cell(row=row, column=8, value="N/F")
            _style_data(ws.cell(row=row, column=8), bg=bg, align="center")

        # Weight
        ws.cell(row=row, column=9, value=landed.get("weight_lbs", "?") if landed else "?")
        _style_data(ws.cell(row=row, column=9), bg=bg, align="center")

        # Courier cost
        if landed:
            ws.cell(row=row, column=10, value=landed.get("courier_cost_dop", 0))
            _style_data(ws.cell(row=row, column=10), bg=bg, align="center")
            ws.cell(row=row, column=10).number_format = '#,##0'
        else:
            ws.cell(row=row, column=10, value="N/F")
            _style_data(ws.cell(row=row, column=10), bg=bg, align="center")

        # Total landed DOP (× qty)
        landed_dop_total = round(landed_dop * qty, 2) if landed_dop > 0 else 0
        if landed_dop_total > 0:
            ws.cell(row=row, column=11, value=landed_dop_total)
            _style_data(ws.cell(row=row, column=11), bg=bg, align="center", bold=True)
            ws.cell(row=row, column=11).number_format = '#,##0'
        else:
            ws.cell(row=row, column=11, value="N/F")
            _style_data(ws.cell(row=row, column=11), bg=YELLOW_BG, align="center")

        # Update running totals with qty-adjusted landed
        if landed_dop_total > 0:
            total_landed += landed_dop_total
            parts_found += 1

        # Delivery days (col 12) — eBay only; RockAuto has no estimate
        if best:
            d_min = best.get("delivery_days_min")
            d_max = best.get("delivery_days_max")
            if d_min is not None and d_max is not None:
                delivery_str = f"{d_min}-{d_max} días"
                any_delivery_known = True
                max_delivery_days = max(max_delivery_days, d_max)
            elif d_max is not None:
                delivery_str = f"≤{d_max} días"
                any_delivery_known = True
                max_delivery_days = max(max_delivery_days, d_max)
            elif d_min is not None:
                delivery_str = f"≥{d_min} días"
                any_delivery_known = True
                max_delivery_days = max(max_delivery_days, d_min)
            else:
                delivery_str = "?"
        else:
            delivery_str = "N/F"
        del_cell = ws.cell(row=row, column=12, value=delivery_str)
        _style_data(del_cell, bg=bg, align="center")

        # Confianza (col 13)
        conf_label = _confidence_label(result, (i - 1) in flagged_indices)
        conf_cell = ws.cell(row=row, column=13, value=conf_label)
        conf_bg = (GREEN_BG if "✅" in conf_label or "🟢" in conf_label
                   else YELLOW_BG if "🟡" in conf_label else ORANGE_BG)
        _style_data(conf_cell, bg=conf_bg, align="center")

        # Source (col 14)
        ws.cell(row=row, column=14, value=best.get("source", "N/F"))
        _style_data(ws.cell(row=row, column=14), bg=bg, align="center")

        # Link (col 15) — HYPERLINK formula for mobile compatibility
        link_url = best.get("url", "")
        if link_url:
            safe_url = link_url.replace('"', '%22')
            link_cell = ws.cell(row=row, column=15,
                                value=f'=HYPERLINK("{safe_url}","Ver →")')
            link_cell.font = LINK_FONT
            link_cell.border = BORDER
            link_cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws.cell(row=row, column=15, value="N/F")
            _style_data(ws.cell(row=row, column=15), bg=bg, align="center")

        ws.row_dimensions[row].height = 20

    # ── Summary row ──
    # Column layout: 1=#, 2=Pieza(orig), 3=Pieza(EN), 4=Cant, 5=Lado,
    #   6=OEM#, 7=USD, 8=Envio, 9=Peso, 10=Courier, 11=Total,
    #   12=Entrega, 13=Confianza, 14=Fuente, 15=Link
    sr = len(results) + 5
    ws.cell(row=sr, column=6, value="TOTALES:").font = Font(bold=True, size=12, name="Calibri")
    _style_data(ws.cell(row=sr, column=6), bold=True)

    if total_landed > 0:
        ws.cell(row=sr, column=11, value=round(total_landed, 0))
        _style_data(ws.cell(row=sr, column=11), bold=True, align="center")
        ws.cell(row=sr, column=11).number_format = '#,##0'

    # Supplier total row (from OCR) — one row below the landed total
    sr_sup = sr + 1
    if supplier_total_dop and supplier_total_dop > 0:
        ws.cell(row=sr_sup, column=6, value="COTIZACIÓN SUPLIDOR:")
        _style_data(ws.cell(row=sr_sup, column=6), bold=True, bg=ORANGE_BG)
        sc2 = ws.cell(row=sr_sup, column=11, value=round(supplier_total_dop, 0))
        sc2.font = Font(bold=True, size=10, name="Calibri")
        sc2.fill = PatternFill("solid", fgColor=ORANGE_BG)
        sc2.border = BORDER
        sc2.alignment = Alignment(horizontal="center", vertical="center")
        sc2.number_format = '#,##0'

    # Stats row
    sr2 = sr_sup + 1
    ws.cell(row=sr2, column=6, value=f"Encontradas: {parts_found}/{len(results)} piezas")
    _style_data(ws.cell(row=sr2, column=6), bg=LIGHT_GRAY)

    # Delivery benchmark row
    sr_del = sr2 + 1
    if any_delivery_known and max_delivery_days > 0:
        delivery_benchmark = f"📦 Todas las piezas llegan en: {max_delivery_days} días (estimado eBay)"
    else:
        delivery_benchmark = "📦 Tiempo de entrega: no disponible (eBay no reportó estimados)"
    ws.merge_cells(f"F{sr_del}:L{sr_del}")
    del_bench_cell = ws.cell(row=sr_del, column=6, value=delivery_benchmark)
    del_bench_cell.font = Font(bold=any_delivery_known, size=10, name="Calibri",
                               color="003DA5" if any_delivery_known else "666666")
    del_bench_cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE if any_delivery_known else LIGHT_GRAY)
    del_bench_cell.border = BORDER
    del_bench_cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Notes ──
    sr3 = sr + 4  # shifted down one row for delivery benchmark
    notes = [
        "Piezas de carroceria se envian sin pintar (primed). Pintura no incluida.",
        f"Conversion usada: RD${exchange_rate:.0f} = $1 USD. Verificar tasa actual.",
        f"Courier: ClickPack RD${clickpack_rate:.0f}/lb. Verificar tarifa vigente.",
        "Verificar compatibilidad con trim/motor especifico antes de ordenar.",
    ]
    for j, note in enumerate(notes):
        ws.cell(row=sr3 + j, column=2, value=note)
        ws.cell(row=sr3 + j, column=2).font = Font(size=9, color="666666", name="Calibri")

    # ── Escenario % sheet ──
    _build_scenario_sheet(wb, vehicle_info, sr, sr_sup, total_landed, supplier_total_dop)

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
